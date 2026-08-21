import csv
import io
import json
from collections import defaultdict
from datetime import UTC, date, datetime, time, timedelta
from pathlib import Path
from typing import Any
from uuid import UUID

import asyncpg
from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response, status

from .analytics import build_productivity_report, validate_range
from .auth import CurrentUser, require_permission, visible_employee_ids
from .database import connection
from .reporting_schemas import (
    ReportExportRequest,
    ReportPresetCreate,
    ReportPresetItem,
    ReportRunItem,
    ReportScheduleCreate,
    ReportScheduleItem,
    ReportTable,
)


router = APIRouter(prefix="/api/v1", tags=["reports"])


async def db() -> asyncpg.Connection:
    async for conn in connection():
        yield conn


def json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime, UUID)):
        return value.isoformat() if hasattr(value, "isoformat") else str(value)
    if isinstance(value, (list, tuple)):
        return [json_value(item) for item in value]
    if isinstance(value, dict):
        return {key: json_value(item) for key, item in value.items()}
    return value


async def productivity_table(
    conn: asyncpg.Connection, user: CurrentUser, start: datetime, end: datetime,
    department_id: UUID | None = None, employee_id: UUID | None = None, basis: str = "planned",
) -> ReportTable:
    report = await build_productivity_report(conn, user, start, end, basis, department_id, employee_id, "employee", "asc")
    columns = ["employee_name", "department_name", "planned_hours", "online_hours", "productive_hours",
               "neutral_hours", "unproductive_hours", "idle_hours", "productive_percent", "delta_pp", "grade"]
    rows = [{
        "employee_id": str(row.employee_id), "employee_name": row.employee_name,
        "department_name": row.department_name or "", "planned_hours": round(row.planned_seconds / 3600, 2),
        "online_hours": round(row.online_seconds / 3600, 2), "productive_hours": round(row.productive_seconds / 3600, 2),
        "neutral_hours": round(row.neutral_seconds / 3600, 2), "unproductive_hours": round(row.unproductive_seconds / 3600, 2),
        "idle_hours": round(row.idle_seconds / 3600, 2), "productive_percent": row.productive_percent,
        "delta_pp": row.delta_productive_pp, "grade": row.grade_label,
    } for row in report.rows]
    return ReportTable(code="productivity", title="Продуктивность сотрудников", columns=columns, rows=rows,
                       summary=json_value(report.totals.model_dump()))


async def departments_table(
    conn: asyncpg.Connection, user: CurrentUser, start: datetime, end: datetime,
) -> ReportTable:
    source = await productivity_table(conn, user, start, end, basis="active")
    grouped: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    counts: dict[str, int] = defaultdict(int)
    for row in source.rows:
        name = row["department_name"] or "Без отдела"; counts[name] += 1
        for key in ("online_hours", "productive_hours", "neutral_hours", "unproductive_hours", "idle_hours"):
            grouped[name][key] += float(row[key])
    rows = []
    for name, values in grouped.items():
        active = values["productive_hours"] + values["neutral_hours"] + values["unproductive_hours"]
        rows.append({"department_name": name, "employees": counts[name], **{key: round(value, 2) for key, value in values.items()},
                     "productive_percent": round(values["productive_hours"] * 100 / active, 1) if active else 0})
    rows.sort(key=lambda item: item["productive_percent"], reverse=True)
    for index, row in enumerate(rows, 1): row["rank"] = index
    columns = ["rank", "department_name", "employees", "online_hours", "productive_hours", "unproductive_hours", "idle_hours", "productive_percent"]
    return ReportTable(code="departments", title="Сравнение отделов", columns=columns, rows=rows)


async def timesheet_table(
    conn: asyncpg.Connection, user: CurrentUser, month: str, department_id: UUID | None = None,
) -> ReportTable:
    try:
        first = date.fromisoformat(month + "-01")
    except ValueError as error:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Месяц должен быть YYYY-MM") from error
    next_month = (first.replace(day=28) + timedelta(days=4)).replace(day=1)
    visible = await visible_employee_ids(conn, user)
    employees = await conn.fetch(
        """SELECT id,full_name,department_name,planned_daily_minutes FROM employees
           WHERE status='active' AND ($1::uuid IS NULL OR department_id=$1)
             AND ($2::uuid[] IS NULL OR id=ANY($2::uuid[])) ORDER BY full_name""",
        department_id, None if visible is None else list(visible),
    )
    activity = await conn.fetch(
        """SELECT employee_id,(ts_start AT TIME ZONE 'UTC')::date AS day,sum(duration_sec)::bigint AS seconds
           FROM activity_events WHERE ts_start<$2 AND ts_end>=$1 AND employee_id=ANY($3::uuid[])
             AND state NOT IN ('LOCKED','BREAK') GROUP BY employee_id,day""",
        datetime.combine(first, time.min, UTC), datetime.combine(next_month, time.min, UTC), [row["id"] for row in employees],
    ) if employees else []
    absences = await conn.fetch(
        """SELECT a.employee_id,a.date_from,a.date_to,t.code FROM absences a JOIN absence_types t ON t.id=a.type_id
           WHERE a.status='approved' AND a.date_from<$2 AND a.date_to>=$1 AND a.employee_id=ANY($3::uuid[])""",
        first, next_month, [row["id"] for row in employees],
    ) if employees else []
    activity_map = {(row["employee_id"], row["day"]): int(row["seconds"]) for row in activity}
    rows = []
    day_count = (next_month - first).days
    for employee in employees:
        cells: dict[str, Any] = {}; fact = 0; plan = 0
        for offset in range(day_count):
            day = first + timedelta(days=offset); code = None
            for absence in absences:
                if absence["employee_id"] == employee["id"] and absence["date_from"] <= day <= absence["date_to"]:
                    code = absence["code"]; break
            seconds = activity_map.get((employee["id"], day), 0); fact += seconds
            if day.isoweekday() <= 5 and code not in {"VACATION", "VACATION_UNPAID", "SICK_LEAVE", "DAY_OFF"}:
                plan += employee["planned_daily_minutes"] * 60
            cells[f"d{day.day}"] = code or round(seconds / 3600, 1) or "—"
        rows.append({"employee_id": str(employee["id"]), "employee_name": employee["full_name"],
                     "department_name": employee["department_name"] or "", "plan_hours": round(plan / 3600, 1),
                     "fact_hours": round(fact / 3600, 1), "deviation_hours": round((fact - plan) / 3600, 1), **cells})
    columns = ["employee_name", "department_name", "plan_hours", "fact_hours", "deviation_hours"] + [f"d{day}" for day in range(1, day_count + 1)]
    return ReportTable(code="timesheet", title=f"Табель за {month}", columns=columns, rows=rows,
                       summary={"legend": {"VACATION": "ОТ", "SICK_LEAVE": "Б", "DAY_OFF": "ОВ"}})


async def apps_table(
    conn: asyncpg.Connection, user: CurrentUser, start: datetime, end: datetime,
    only_unproductive: bool = False, employee_id: UUID | None = None,
) -> ReportTable:
    validate_range(start, end); visible = await visible_employee_ids(conn, user)
    rows = await conn.fetch(
        """SELECT e.id AS employee_id,e.full_name AS employee_name,COALESCE(d.name,e.department_name) AS department_name,
                  CASE WHEN a.url_domain IS NOT NULL AND a.url_domain<>'' THEN 'site' ELSE 'application' END AS kind,
                  CASE WHEN a.url_domain IS NOT NULL AND a.url_domain<>'' THEN a.url_domain ELSE COALESCE(a.app_name,a.process_name,'Неизвестно') END AS item,
                  COALESCE(c.name,'Без категории') AS category,COALESCE(c.productivity,a.state) AS productivity,
                  sum(EXTRACT(EPOCH FROM (LEAST(a.ts_end,$2)-GREATEST(a.ts_start,$1))))::bigint AS seconds
           FROM activity_events a JOIN employees e ON e.id=a.employee_id LEFT JOIN departments d ON d.id=e.department_id
           LEFT JOIN categories c ON c.id=a.category_id WHERE a.ts_start<$2 AND a.ts_end>$1
             AND a.state IN ('PRODUCTIVE','NEUTRAL','UNPRODUCTIVE')
             AND ($3::boolean=false OR COALESCE(c.productivity,a.state)='UNPRODUCTIVE')
             AND ($4::uuid IS NULL OR e.id=$4) AND ($5::uuid[] IS NULL OR e.id=ANY($5::uuid[]))
           GROUP BY e.id,d.name,kind,item,category,productivity ORDER BY seconds DESC""",
        start, end, only_unproductive, employee_id, None if visible is None else list(visible),
    )
    result = [{**json_value(dict(row)), "hours": round(int(row["seconds"]) / 3600, 2)} for row in rows]
    columns = ["employee_name", "department_name", "kind", "item", "category", "productivity", "hours"]
    return ReportTable(code="apps", title="Приложения и сайты" + (" — топ непродуктивного" if only_unproductive else ""), columns=columns, rows=result)


async def discipline_table(
    conn: asyncpg.Connection, user: CurrentUser, start: date, end: date,
) -> ReportTable:
    visible = await visible_employee_ids(conn, user)
    rows = await conn.fetch(
        """SELECT e.id AS employee_id,e.full_name AS employee_name,COALESCE(d.name,e.department_name) AS department_name,
                  count(*) FILTER(WHERE t.code='LATE_VALID')::int AS late_valid,
                  count(*) FILTER(WHERE t.code='LATE_INVALID')::int AS late_invalid,
                  COALESCE(sum(a.minutes) FILTER(WHERE t.code LIKE 'LATE_%'),0)::int AS late_minutes,
                  count(*) FILTER(WHERE t.code='EARLY_LEAVE')::int AS early_leaves,
                  count(*) FILTER(WHERE t.code='ABSENCE_UNEXCUSED')::int AS absences_unexcused,
                  count(*) FILTER(WHERE t.code='VIOLATION')::int AS violations,
                  COALESCE(sum(a.severity) FILTER(WHERE t.code='VIOLATION'),0)::int AS severity_points
           FROM employees e LEFT JOIN departments d ON d.id=e.department_id
           LEFT JOIN absences a ON a.employee_id=e.id AND a.date_from<=$2 AND a.date_to>=$1 AND a.status='approved'
           LEFT JOIN absence_types t ON t.id=a.type_id
           WHERE ($3::uuid[] IS NULL OR e.id=ANY($3::uuid[])) GROUP BY e.id,d.name ORDER BY violations DESC,late_minutes DESC""",
        start, end, None if visible is None else list(visible),
    )
    columns = ["employee_name", "department_name", "late_valid", "late_invalid", "late_minutes", "early_leaves", "absences_unexcused", "violations", "severity_points"]
    return ReportTable(code="discipline", title="Дисциплина", columns=columns, rows=[json_value(dict(row)) for row in rows])


async def absences_table(
    conn: asyncpg.Connection, user: CurrentUser, start: date, end: date,
) -> ReportTable:
    visible = await visible_employee_ids(conn, user)
    rows = await conn.fetch(
        """SELECT e.id AS employee_id,e.full_name AS employee_name,COALESCE(d.name,e.department_name) AS department_name,
                  t.code,t.name AS type_name,a.date_from,a.date_to,(a.date_to-a.date_from+1)::int AS days,
                  a.status,a.reason,a.is_auto,COALESCE(v.opening_days+v.accrued_days-v.used_days,0) AS vacation_balance
           FROM absences a JOIN employees e ON e.id=a.employee_id JOIN absence_types t ON t.id=a.type_id
           LEFT JOIN departments d ON d.id=e.department_id LEFT JOIN vacation_balances v ON v.employee_id=e.id AND v.balance_year=EXTRACT(YEAR FROM $1::date)
           WHERE a.date_from<=$2 AND a.date_to>=$1 AND ($3::uuid[] IS NULL OR e.id=ANY($3::uuid[]))
           ORDER BY a.date_from,e.full_name""", start, end, None if visible is None else list(visible),
    )
    columns = ["employee_name", "department_name", "type_name", "date_from", "date_to", "days", "status", "reason", "vacation_balance"]
    return ReportTable(code="absences", title="Отсутствия", columns=columns, rows=[json_value(dict(row)) for row in rows])


async def daily_activity_table(
    conn: asyncpg.Connection, user: CurrentUser, employee_id: UUID, day: date,
) -> ReportTable:
    visible = await visible_employee_ids(conn, user)
    if visible is not None and employee_id not in visible:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Сотрудник не найден")
    start = datetime.combine(day, time.min, UTC); end = start + timedelta(days=1)
    rows = await conn.fetch(
        """SELECT a.ts_start,a.ts_end,a.duration_sec,a.state,COALESCE(a.app_name,a.process_name) AS application,
                  a.window_title,a.url_domain,a.url_path,c.name AS category,
                  (SELECT s.id FROM screenshots s WHERE s.activity_event_id=a.id ORDER BY s.taken_at LIMIT 1) AS screenshot_id,
                  (SELECT ss.id FROM stream_sessions ss WHERE ss.employee_id=a.employee_id AND ss.started_at<=a.ts_start
                   AND COALESCE(ss.ended_at,now())>=a.ts_start ORDER BY ss.started_at DESC LIMIT 1) AS stream_session_id
           FROM activity_events a LEFT JOIN categories c ON c.id=a.category_id
           WHERE a.employee_id=$1 AND a.ts_start<$3 AND a.ts_end>$2 ORDER BY a.ts_start""",
        employee_id, start, end,
    )
    result = []
    for row in rows:
        item = json_value(dict(row))
        item["screenshot_url"] = f"/api/v1/screenshots/{row['screenshot_id']}/image" if row["screenshot_id"] else None
        result.append(item)
    columns = ["ts_start", "ts_end", "duration_sec", "state", "application", "window_title", "url_domain", "category", "screenshot_url"]
    return ReportTable(code="daily_activity", title=f"Активность за {day}", columns=columns, rows=result)


async def anomalies_table(
    conn: asyncpg.Connection, user: CurrentUser, start: date, end: date, drop_pp: float = 20,
) -> ReportTable:
    visible = await visible_employee_ids(conn, user)
    rows = await conn.fetch(
        """WITH daily AS (
             SELECT employee_id,(ts_start AT TIME ZONE 'UTC')::date AS day,
                    100.0*sum(duration_sec) FILTER(WHERE state='PRODUCTIVE') /
                    NULLIF(sum(duration_sec) FILTER(WHERE state IN ('PRODUCTIVE','NEUTRAL','UNPRODUCTIVE')),0) AS pct
             FROM activity_events WHERE ts_start >= $1::date - interval '14 days' AND ts_start < $2::date + interval '1 day'
             GROUP BY employee_id,day
           ), scored AS (
             SELECT employee_id,day,pct,avg(pct) OVER(PARTITION BY employee_id ORDER BY day ROWS BETWEEN 14 PRECEDING AND 1 PRECEDING) AS baseline
             FROM daily
           )
           SELECT s.employee_id,e.full_name AS employee_name,e.department_name,s.day,round(s.pct,1) AS productive_percent,
                  round(s.baseline,1) AS baseline_percent,round(s.baseline-s.pct,1) AS drop_pp
           FROM scored s JOIN employees e ON e.id=s.employee_id WHERE s.day BETWEEN $1 AND $2
             AND s.baseline-s.pct >= $3 AND ($4::uuid[] IS NULL OR s.employee_id=ANY($4::uuid[]))
           ORDER BY drop_pp DESC""",
        start, end, drop_pp, None if visible is None else list(visible),
    )
    columns = ["employee_name", "department_name", "day", "productive_percent", "baseline_percent", "drop_pp"]
    return ReportTable(code="anomalies", title="Аномалии продуктивности", columns=columns, rows=[json_value(dict(row)) for row in rows], summary={"drop_threshold_pp": drop_pp})


@router.get("/reports/departments", response_model=ReportTable)
async def report_departments(range_start: datetime = Query(alias="from"), range_end: datetime = Query(alias="to"),
                             conn: asyncpg.Connection = Depends(db), user: CurrentUser = Depends(require_permission("reports:view"))) -> ReportTable:
    validate_range(range_start, range_end); return await departments_table(conn, user, range_start, range_end)


@router.get("/reports/timesheet", response_model=ReportTable)
async def report_timesheet(month: str, department_id: UUID | None = None, conn: asyncpg.Connection = Depends(db),
                           user: CurrentUser = Depends(require_permission("reports:view"))) -> ReportTable:
    return await timesheet_table(conn, user, month, department_id)


@router.get("/reports/apps", response_model=ReportTable)
async def report_apps(range_start: datetime = Query(alias="from"), range_end: datetime = Query(alias="to"), only_unproductive: bool = False,
                      employee_id: UUID | None = None, conn: asyncpg.Connection = Depends(db),
                      user: CurrentUser = Depends(require_permission("reports:view"))) -> ReportTable:
    return await apps_table(conn, user, range_start, range_end, only_unproductive, employee_id)


@router.get("/reports/discipline", response_model=ReportTable)
async def report_discipline(range_start: date = Query(alias="from"), range_end: date = Query(alias="to"), conn: asyncpg.Connection = Depends(db),
                            user: CurrentUser = Depends(require_permission("reports:view"))) -> ReportTable:
    return await discipline_table(conn, user, range_start, range_end)


@router.get("/reports/absences", response_model=ReportTable)
async def report_absences(range_start: date = Query(alias="from"), range_end: date = Query(alias="to"), conn: asyncpg.Connection = Depends(db),
                          user: CurrentUser = Depends(require_permission("reports:view"))) -> ReportTable:
    return await absences_table(conn, user, range_start, range_end)


@router.get("/reports/daily-activity", response_model=ReportTable)
async def report_daily_activity(employee_id: UUID, day: date, conn: asyncpg.Connection = Depends(db),
                                user: CurrentUser = Depends(require_permission("reports:view"))) -> ReportTable:
    return await daily_activity_table(conn, user, employee_id, day)


@router.get("/reports/anomalies", response_model=ReportTable)
async def report_anomalies(range_start: date = Query(alias="from"), range_end: date = Query(alias="to"), drop_pp: float = Query(default=20, ge=1, le=100),
                           conn: asyncpg.Connection = Depends(db), user: CurrentUser = Depends(require_permission("reports:view"))) -> ReportTable:
    return await anomalies_table(conn, user, range_start, range_end, drop_pp)


def cell_text(value: Any) -> str:
    if value is None: return ""
    if isinstance(value, bool): return "Да" if value else "Нет"
    if isinstance(value, (dict, list)): return json.dumps(value, ensure_ascii=False)
    return str(value)


def export_csv(table: ReportTable, columns: list[str]) -> bytes:
    output = io.StringIO(); writer = csv.writer(output, delimiter=";"); writer.writerow(columns)
    for row in table.rows: writer.writerow([cell_text(row.get(column)) for column in columns])
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def export_xlsx(table: ReportTable, columns: list[str]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    workbook = Workbook(); sheet = workbook.active; sheet.title = table.code[:31]
    sheet.append(columns)
    for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="294B36")
    for row in table.rows:
        sheet.append([cell_text(row.get(column)) for column in columns])
        if "productive_percent" in columns:
            value = float(row.get("productive_percent") or 0); index = columns.index("productive_percent") + 1
            sheet.cell(sheet.max_row, index).fill = PatternFill("solid", fgColor="C6EFCE" if value >= 75 else "FFEB9C" if value >= 50 else "FFC7CE")
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        sheet.column_dimensions[letter].width = min(45, max(10, max(len(cell_text(cell.value)) for cell in column_cells) + 2))
        for cell in column_cells: cell.alignment = Alignment(vertical="top", wrap_text=True)
    output = io.BytesIO(); workbook.save(output); return output.getvalue()


def export_pdf(table: ReportTable, columns: list[str]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    font_path = next((path for path in [Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"), Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf")] if path.exists()), None)
    font = "Helvetica"
    if font_path:
        pdfmetrics.registerFont(TTFont("WorkforceUnicode", str(font_path))); font = "WorkforceUnicode"
    output = io.BytesIO(); document = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet(); styles["Title"].fontName = font; styles["BodyText"].fontName = font; styles["BodyText"].fontSize = 6
    selected = columns[:14]
    data = [[Paragraph(column, styles["BodyText"]) for column in selected]]
    for row in table.rows[:500]: data.append([Paragraph(cell_text(row.get(column))[:500], styles["BodyText"]) for column in selected])
    grid = Table(data, repeatRows=1); grid.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), font), ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#294B36")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), .25, colors.HexColor("#D6DDD7")),
        ("VALIGN", (0,0), (-1,-1), "TOP"), ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F3F6F2")]),
    ]))
    document.build([Paragraph(table.title, styles["Title"]), Spacer(1, 10), grid]); return output.getvalue()


async def table_from_export(conn: asyncpg.Connection, user: CurrentUser, code: str, filters: dict) -> ReportTable:
    try:
        if code == "productivity": return await productivity_table(conn, user, datetime.fromisoformat(filters["from"]), datetime.fromisoformat(filters["to"]), basis=filters.get("basis", "planned"))
        if code == "departments": return await departments_table(conn, user, datetime.fromisoformat(filters["from"]), datetime.fromisoformat(filters["to"]))
        if code == "timesheet": return await timesheet_table(conn, user, filters["month"])
        if code == "apps": return await apps_table(conn, user, datetime.fromisoformat(filters["from"]), datetime.fromisoformat(filters["to"]), bool(filters.get("only_unproductive")))
        if code == "discipline": return await discipline_table(conn, user, date.fromisoformat(filters["from"]), date.fromisoformat(filters["to"]))
        if code == "absences": return await absences_table(conn, user, date.fromisoformat(filters["from"]), date.fromisoformat(filters["to"]))
        if code == "daily_activity": return await daily_activity_table(conn, user, UUID(filters["employee_id"]), date.fromisoformat(filters["day"]))
        if code == "anomalies": return await anomalies_table(conn, user, date.fromisoformat(filters["from"]), date.fromisoformat(filters["to"]), float(filters.get("drop_pp", 20)))
    except (KeyError, ValueError) as error:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, f"Некорректные фильтры: {error}") from error
    raise HTTPException(status.HTTP_404_NOT_FOUND, "Неизвестный отчёт")


@router.post("/reports/{code}/export")
async def export_report(code: str, payload: ReportExportRequest, request: Request, conn: asyncpg.Connection = Depends(db),
                        user: CurrentUser = Depends(require_permission("reports:view"))) -> Response:
    table = await table_from_export(conn, user, code, payload.filters); columns = payload.columns or table.columns
    if any(column not in table.columns for column in columns):
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Неизвестный столбец")
    if payload.format == "csv": content, media = export_csv(table, columns), "text/csv; charset=utf-8"
    elif payload.format == "xlsx": content, media = export_xlsx(table, columns), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else: content, media = export_pdf(table, columns), "application/pdf"
    ip = request.headers.get("x-forwarded-for", "").split(",", 1)[0] or (request.client.host if request.client else None)
    await conn.execute("""INSERT INTO audit_log(user_id,action,object_type,object_id,details_json)
                          VALUES($1,'report_exported','report',$2,$3::jsonb)""", user.id, code,
                       json.dumps({"format": payload.format, "rows": len(table.rows), "ip": ip}))
    return Response(content=content, media_type=media, headers={"Content-Disposition": f'attachment; filename="{code}.{payload.format}"'})


@router.get("/report-presets", response_model=list[ReportPresetItem])
async def list_presets(conn: asyncpg.Connection = Depends(db), user: CurrentUser = Depends(require_permission("reports:view"))) -> list[ReportPresetItem]:
    rows = await conn.fetch("""SELECT id,name,report_code,filters_json AS filters,columns_json AS columns,created_at,updated_at
                              FROM saved_report_presets WHERE user_id=$1 ORDER BY updated_at DESC""", user.id)
    return [ReportPresetItem(**dict(row)) for row in rows]


@router.post("/report-presets", response_model=ReportPresetItem, status_code=201)
async def create_preset(payload: ReportPresetCreate, conn: asyncpg.Connection = Depends(db),
                        user: CurrentUser = Depends(require_permission("reports:view"))) -> ReportPresetItem:
    row = await conn.fetchrow("""INSERT INTO saved_report_presets(user_id,name,report_code,filters_json,columns_json)
                                 VALUES($1,$2,$3,$4::jsonb,$5::jsonb)
                                 RETURNING id,name,report_code,filters_json AS filters,columns_json AS columns,created_at,updated_at""",
                              user.id, payload.name, payload.report_code, json.dumps(payload.filters), json.dumps(payload.columns) if payload.columns else None)
    return ReportPresetItem(**dict(row))


@router.delete("/report-presets/{preset_id}", status_code=204)
async def delete_preset(preset_id: UUID, conn: asyncpg.Connection = Depends(db),
                        user: CurrentUser = Depends(require_permission("reports:view"))) -> None:
    result = await conn.execute("DELETE FROM saved_report_presets WHERE id=$1 AND user_id=$2", preset_id, user.id)
    if result.endswith(" 0"): raise HTTPException(status.HTTP_404_NOT_FOUND, "Пресет не найден")


@router.get("/report-schedules", response_model=list[ReportScheduleItem])
async def list_report_schedules(conn: asyncpg.Connection = Depends(db), user: CurrentUser = Depends(require_permission("reports:view"))) -> list[ReportScheduleItem]:
    rows = await conn.fetch("""SELECT id,report_code,filters_json AS filters,recipients,cron,format,enabled,last_run_at,next_run_at,created_at
                              FROM report_schedules WHERE user_id=$1 ORDER BY created_at DESC""", user.id)
    return [ReportScheduleItem(**dict(row)) for row in rows]


@router.post("/report-schedules", response_model=ReportScheduleItem, status_code=201)
async def create_report_schedule(payload: ReportScheduleCreate, conn: asyncpg.Connection = Depends(db),
                                 user: CurrentUser = Depends(require_permission("reports:view"))) -> ReportScheduleItem:
    row = await conn.fetchrow("""INSERT INTO report_schedules(user_id,report_code,filters_json,recipients,cron,format,enabled,next_run_at)
                                 VALUES($1,$2,$3::jsonb,$4,$5,$6,$7,now())
                                 RETURNING id,report_code,filters_json AS filters,recipients,cron,format,enabled,last_run_at,next_run_at,created_at""",
                              user.id, payload.report_code, json.dumps(payload.filters), payload.recipients, payload.cron, payload.format, payload.enabled)
    return ReportScheduleItem(**dict(row))


@router.delete("/report-schedules/{schedule_id}", status_code=204)
async def delete_report_schedule(schedule_id: UUID, conn: asyncpg.Connection = Depends(db),
                                 user: CurrentUser = Depends(require_permission("reports:view"))) -> None:
    result = await conn.execute("DELETE FROM report_schedules WHERE id=$1 AND user_id=$2", schedule_id, user.id)
    if result.endswith(" 0"): raise HTTPException(status.HTTP_404_NOT_FOUND, "Расписание не найдено")


@router.get("/report-runs", response_model=list[ReportRunItem])
async def list_report_runs(conn: asyncpg.Connection = Depends(db), user: CurrentUser = Depends(require_permission("reports:view"))) -> list[ReportRunItem]:
    rows = await conn.fetch("""SELECT r.id,r.schedule_id,r.report_code,r.status,r.recipients,r.storage_key,r.error,r.created_at,r.finished_at
                              FROM report_runs r LEFT JOIN report_schedules s ON s.id=r.schedule_id
                              WHERE s.user_id=$1 OR r.schedule_id IS NULL ORDER BY r.created_at DESC LIMIT 200""", user.id)
    return [ReportRunItem(**dict(row)) for row in rows]
